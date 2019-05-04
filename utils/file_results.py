import pickle
from openpyxl import Workbook

def xlsfile(Param):

    # load results
    from statistics import mean
    import csv

    dict_data = pickle.load( open( "savedata/data_save.p", "rb" ) )
    dict_results = pickle.load( open( "savedata/results_save.p", "rb" ) )
    dict_hydro = pickle.load( open( "savedata/hydro_save.p", "rb" ) )
    dict_format = pickle.load( open( "savedata/format_save.p", "rb" ) )
    dict_lines = pickle.load( open( "savedata/lines_save.p", "rb" ) )
    dict_sim = pickle.load( open( "savedata/format_sim_save.p", "rb" ) )

#    dict_windenergy = pickle.load( open( "savedata/windspeed_save.p", "rb" ) )

    # dictionary
    genThermal = dict_results["genThermal"]
    genSmall = dict_results["genSmall"]
    genHydro = dict_results["genHydro"]
    genBatteries = dict_results["genBatteries"]
    genDeficit = dict_results["genDeficit"]
    loadBatteries = dict_results["loadBatteries"]
    lvlBatteries = dict_results["lvlBatteries"]
    lvlHydro = dict_results["lvlHydro"]
    linTransfer = dict_results["linTransfer"]
    spillHydro = dict_results["spillHydro"]
    genwind = dict_results["genwind"]
    genRnws = dict_results["genRnws"]
    marg_costs = dict_results["marg_costs"]
    emsscurve = dict_results["emsscurve"]
    l_limits = dict_lines["l_limits"]

    batteries = dict_data['batteries']
    hydroPlants = dict_data['hydroPlants']
    thermalPlants = dict_data['thermalPlants']
    smallPlants = dict_data['smallPlants']
    numBlocks = dict_format['numBlocks']
    df_demand = dict_format['demand']
    numAreas = dict_data['numAreas']
    battData = dict_data['battData']
    windPlants = dict_data['windPlants']
    volData = dict_data['volData']
    prodFactor = dict_hydro['prodFactor']
    thermalData = dict_data['thermalData']
    smallData = dict_data['smallData']
    linesData = dict_data['linesData']
#    costData = dict_data['costData']
    areasData = dict_data['areasData']
    yearvector = dict_sim['yearvector']

    #lenblk = range(numBlocks); lenthm = range(len(thermalPlants))
    scenarios = Param.seriesForw
    stages = Param.stages-Param.bnd_stages
    
    lenstg = range(stages); lensc = range(scenarios)
#    lenareas = range(numAreas)

    # Create file
    wb = Workbook()
    dest_filename = 'results/General_results.xlsx'

    ws0 = wb.active
    ws0.title = "Summary"

    ws0['A2'] = 'Scenarios'; ws0['B2'] = scenarios
    ws0['A3'] = 'Stages'; ws0['B3'] = stages
    ws0['A4'] = 'Blocks'; ws0['B4'] = numBlocks
    ws0['A5'] = 'Areas'; ws0['B5'] = numAreas
    ws0['A6'] = 'Hydro plants'; ws0['B6'] = len(hydroPlants)
    ws0['A7'] = 'Small plants'; ws0['B7'] = len(thermalPlants)
    ws0['A8'] = 'Thermal plants'; ws0['B8'] = len(smallPlants)
    ws0['A9'] = 'Wind plants'; ws0['B9'] = len(windPlants)
    ws0['A10'] = 'Batteries'; ws0['B10'] = len(batteries)
    ws0['A11'] = 'LinksAreas'; ws0['B11'] = len(linesData)

    ###########################################################################

    # Hydro generation results
    plants = len(hydroPlants)
    genHFinal = [[[[] for y in range(scenarios) ] for z in range(plants) ] for x in range(stages)]

    for i in range(scenarios):
        for j in range(stages):
           for k in range(plants):
                # amount of energy produced
                valuegen = 0
                for x in range(numBlocks):
                    valuegen += genHydro[i][j][k][x]*prodFactor[k][j]
                genHFinal[j][k][i] = valuegen

    # write file results
    ws1 = wb.create_sheet(title="HydroGen")
    for j in range(stages):
        for k in range(plants):
            valueG = 0
            for i in range(scenarios):
               valueG += genHFinal[j][k][i]

            _ = ws1.cell(column=k+2, row=3+(j+1), value=valueG/scenarios)
            _ = ws1.cell(column=1, row=3+(j+1), value=j+1)
            _ = ws1.cell(column=k+2, row=2, value=hydroPlants[k])
            _ = ws1.cell(column=k+2, row=3, value= 'Area:'+str(volData[11][k]))
    ws1['A3'] = 'Stage'

    # Hydro generation - write csv
    genhydro = [[[] for _ in range(plants*scenarios + 2) ] for _ in range(stages*numBlocks + 2)]
    for i in range(scenarios):
        for j in range(stages):
           for k in range(plants):
                genhydro[0][1+(i+1)+(k+1)*scenarios-scenarios] = hydroPlants[k]
                genhydro[1][1+(i+1)+(k+1)*scenarios-scenarios] = 'Scenario:'+str(i+1)
                for x in range(numBlocks):
                    genhydro[1+(j+1)*numBlocks-numBlocks+(x+1)][1+(i+1)+(k+1)*scenarios-scenarios] = genHydro[i][j][k][x]*prodFactor[k][j]
                    genhydro[1+(j+1)*numBlocks-numBlocks+(x+1)][1] = x+1
                    genhydro[1+(j+1)*numBlocks-numBlocks+(x+1)][0] = j+1
    genhydro[0][0] = 'MWh'
    genhydro[1][0] = 'Stage'
    genhydro[1][1] = 'Block'

    with open("results/csv_variables/HydroGeneration.csv", "w") as f:
        writer = csv.writer(f)
        writer.writerows(genhydro)

    # spill of water - write csv
    spillwater = [[[] for _ in range(plants*scenarios + 2) ] for _ in range(stages*numBlocks + 2)]

    for i in range(scenarios):
        for j in range(stages):
           for k in range(plants):
                spillwater[0][1+(i+1)+(k+1)*scenarios-scenarios] = hydroPlants[k]
                spillwater[1][1+(i+1)+(k+1)*scenarios-scenarios] = 'Scenario:'+str(i+1)
                for x in range(numBlocks):
                    spillwater[1+(j+1)*numBlocks-numBlocks+(x+1)][1+(i+1)+(k+1)*scenarios-scenarios] = spillHydro[i][j][k][x]/(yearvector[j]*3600*1e-6)
                    spillwater[1+(j+1)*numBlocks-numBlocks+(x+1)][1] = x+1
                    spillwater[1+(j+1)*numBlocks-numBlocks+(x+1)][0] = j+1
    spillwater[0][0] = 'spill m3/s'
    spillwater[1][0] = 'Stage'
    spillwater[1][1] = 'Block'

    with open("results/csv_variables/spillHydro.csv", "w") as f:
        writer = csv.writer(f)
        writer.writerows(spillwater)

    # levels of reservoirs at each stage
    lvlhydro = [[[] for _ in range(plants*scenarios + 1) ] for _ in range(stages + 2)]
    for i in range(scenarios):
        for j in range(stages):
           for k in range(plants):
                lvlhydro[0][(i+1)+(k+1)*scenarios-scenarios] = hydroPlants[k]
                lvlhydro[1][(i+1)+(k+1)*scenarios-scenarios] = 'Scenario:'+str(i+1)
                lvlhydro[j+2][(i+1)+(k+1)*scenarios-scenarios] = lvlHydro[i][j][k]
                lvlhydro[j+2][0] = j+1
    lvlhydro[0][0] = 'Hm3'
    lvlhydro[1][0] = 'Stage'

    with open("results/csv_variables/LevelReservoirs.csv", "w") as f:
        writer = csv.writer(f)
        writer.writerows(lvlhydro)

    # level of reservoirs at the end of each stage


    # level of reservoirs csv file
    ws10 = wb.create_sheet(title="LevelHydro")
    for j in range(stages):
        for k in range(plants):
            valueL = 0
            for i in range(scenarios):
                valueL += lvlHydro[i][j][k]
            _ = ws10.cell(column=k+2, row=3+(j+1), value=valueL/scenarios)
            _ = ws10.cell(column=1, row=3+(j+1), value=j+1)
            _ = ws10.cell(column=k+2,row=2, value=hydroPlants[k])
            _ = ws10.cell(column=k+2, row=3, value= 'Area:'+str(volData[11][k]))
    ws10['A3'] = 'Stage'

    # Aggregated reservoir
    levelhydroScn = [[[[] for y in range(plants) ] for z in range(scenarios) ] for x in range(stages)]
    for i in range(scenarios):
        for j in range(stages):
           for k in range(plants):
               levelhydroScn[j][i][k] = lvlHydro[i][j][k]

    Aggrhydro = [[[] for z in range(3) ] for x in range(stages)]
    for j in range(stages):
        sumpl = []
        for i in range(scenarios):
            sumpl.append(sum(levelhydroScn[j][i]))

        Aggrhydro[j][1] = mean(sumpl)
        Aggrhydro[j][0] = min(sumpl); Aggrhydro[j][2] = max(sumpl)

    # write results file
    ws10_2 = wb.create_sheet(title="AggLevelHydro")
    for j in range(stages):
        for k in range(3):
            _ = ws10_2.cell(column=1+(k+1),row=2+(j+1), value=Aggrhydro[j][k])
            _ = ws10_2.cell(column=1, row=2+(j+1), value=j+1)
            _ = ws10_2.cell(column=2,row=2, value='min[Hm3]' )
            _ = ws10_2.cell(column=3,row=2, value='aveg[Hm3]' )
            _ = ws10_2.cell(column=4,row=2, value='max[Hm3]' )
    ws10_2['A2'] = 'Stage'

    ###########################################################################

    # Thermal generation results
    plants = len(thermalPlants)
    genTFinal = [[[[] for y in range(scenarios) ] for z in range(plants) ] for x in range(stages)]

    for i in range(scenarios):
        for j in range(stages):
           for k in range(plants):
                # amount of energy produced
                valuegen = 0
                for x in range(numBlocks):
                    valuegen += genThermal[i][j][k][x]
                genTFinal[j][k][i] = valuegen
    
    # write file results
    ws2 = wb.create_sheet(title="ThermalGen")
    for j in range(stages):
        for k in range(plants):
            valueG = 0
            for i in range(scenarios):
               valueG += genTFinal[j][k][i]

            _ = ws2.cell(column=k+2, row=3+(j+1), value=valueG/scenarios)
            _ = ws2.cell(column=1, row=3+(j+1), value=j+1)
            _ = ws2.cell(column=k+2, row=2, value=thermalPlants[k])
            _ = ws2.cell(column=k+2, row=3, value= 'Area:'+str(thermalData[3][k]))
    ws2['A3'] = 'Stage'

#    # Thermal generation - write csv
#    genhydro = [[[] for _ in range(plants*scenarios + 2) ] for _ in range(stages*numBlocks + 2)]
#    for i in range(scenarios):
#        for j in range(stages):
#           for k in range(plants):
#                genhydro[0][1+(i+1)+(k+1)*scenarios-scenarios] = hydroPlants[k]
#                genhydro[1][1+(i+1)+(k+1)*scenarios-scenarios] = 'Scenario:'+str(i+1)
#                for x in range(numBlocks):
#                    genhydro[1+(j+1)*numBlocks-numBlocks+(x+1)][1+(i+1)+(k+1)*scenarios-scenarios] = genHydro[i][j][k][x]*prodFactor[k][j]
#                    genhydro[1+(j+1)*numBlocks-numBlocks+(x+1)][1] = x+1
#                    genhydro[1+(j+1)*numBlocks-numBlocks+(x+1)][0] = j+1
#    genhydro[0][0] = 'MWh'
#    genhydro[1][0] = 'Stage'
#    genhydro[1][1] = 'Block'


###############################################################################

    # Small plants analysis

    ws22 = wb.create_sheet(title="SmallGen")

    plants = len(smallPlants)
    genSFinal = [[[[] for y in range(scenarios) ] for z in range(plants) ] for x in range(stages)]

    # check if the selected plant in list is actually generating / otherwise cost is zero
    for i in range(scenarios):
        for j in range(stages):
           for k in range(plants):
               # amount of energy produced
               valuegen = 0
               for x in range(numBlocks):

                   valuegen += genSmall[i][j][k][x]

                   # write results
                   _ = ws22.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                row=4+(j+1)*numBlocks-numBlocks+(x+1), value=genSmall[i][j][k][x])
                   _ = ws22.cell(column=2,
                                row=4+(j+1)*numBlocks-numBlocks+(x+1), value=x+1)
                   _ = ws22.cell(column=1,
                                row=4+(j+1)*numBlocks-numBlocks+(x+1), value=j+1)
                   _ = ws22.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                row=2, value=smallPlants[k])
                   _ = ws22.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                row=4, value='Scenario:'+str(i+1) )
                   _ = ws22.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                row=3, value='Area:'+str(int(smallData[3][k])) )
               genSFinal[j][k][i] = valuegen

    ws22['B4'] = 'block'; ws22['A4'] = 'Stage'

    ###########################################################################

    # Marginal cost
    ws21 = wb.create_sheet(title="MarginalCost")
    genMFinal = [[[[] for z in range(scenarios)] for z in range(numAreas) ] for x in range(stages)]

    for k in range(scenarios):
        for i in range(numAreas):
            _ = ws21.cell(column=1+(k+1)+(i+1)*scenarios-scenarios,row=3, value='Scenario:'+str(k+1) )
            _ = ws21.cell(column=1+(k+1)+(i+1)*scenarios-scenarios,row=2, value='Area:'+ areasData[0][i] )
            for j in range(stages):
                valueC = 0
                for x in range(numBlocks):
                    valueC2 = sum(marg_costs[k][j][i][x])
                    if Param.dist_f[0] is True:
                        valueC2 = - sum(marg_costs[k][j][i][x])
                    elif Param.dist_f[1] is True:
                        valueC2 = - sum(marg_costs[k][j][i][x])
                    else:
                        valueC2 = sum(marg_costs[k][j][i][x])
                    if valueC2 > valueC:
                        valueC = valueC2

                genMFinal[j][i][k] = valueC #/numBlocks

                _ = ws21.cell(column=1+(k+1)+(i+1)*scenarios-scenarios,row=3+(j+1), value=valueC)
                _ = ws21.cell(column=1, row=3+(j+1), value=j+1)

    ws21['A3'] = 'Stage'

    ws21_2 = wb.create_sheet(title="MarginalArea")

    for k in range(numAreas):
        _ = ws21_2.cell(column=1+(k+1),row=2, value='Area:'+ areasData[0][k] )
        for j in range(stages):
            _ = ws21_2.cell(column=1+(k+1), row=2+(j+1), value=sum(genMFinal[j][k])/scenarios)
            _ = ws21_2.cell(column=1, row=2+(j+1), value=j+1)

    ws21_2['A2'] = 'Stage'

    ###########################################################################

    # Batteries generation
    ws4 = wb.create_sheet(title="BatteriesGen")

    plants = len(batteries)
    genBFinal = [[[[] for y in range(scenarios) ] for z in range(plants) ] for x in range(stages)]

    for i in range(scenarios):
        for j in range(stages):
           for k in range(plants):
               valueB = 0
               for x in range(numBlocks):
                   valueB += genBatteries[i][j][k][x]

                   _ = ws4.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                row=4+(j+1)*numBlocks-numBlocks+(x+1), value=genBatteries[i][j][k][x])
                   _ = ws4.cell(column=2,
                                row=4+(j+1)*numBlocks-numBlocks+(x+1), value=x+1)
                   _ = ws4.cell(column=1,
                                row=4+(j+1)*numBlocks-numBlocks+(x+1), value=j+1)
                   _ = ws4.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                row=2, value=batteries[k])
                   _ = ws4.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                row=3, value='Area:'+str(int(battData[9][k])) )
                   _ = ws4.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                row=4, value='Scenario:'+str(i+1) )
               genBFinal[j][k][i] = valueB

    ws4['B4'] = 'block'; ws4['A4'] = 'Stage'

    genBBlockFinal = [[0]*numBlocks for x in range(stages)]
    for i in range(scenarios):
        for j in range(stages):
            for k in range(plants):
                for x in range(numBlocks):
                    genBBlockFinal[j][x] += genBatteries[i][j][k][x]
    

    ###########################################################################

    ws5 = wb.create_sheet(title="Deficit")

    genDFinal = [[[[] for y in range(scenarios) ] for z in range(numAreas) ] for x in range(stages)]

    for k in range(numAreas):
        for i in range(scenarios):
            for j in range(stages):
                valueD = 0

                for x in range(numBlocks):
                    valueD += genDeficit[k][i][j][x]

                    _ = ws5.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                 row=3+(j+1)*numBlocks-numBlocks+(x+1),
                                 value=genDeficit[k][i][j][x])
                    _ = ws5.cell(column=2,
                                 row=3+(j+1)*numBlocks-numBlocks+(x+1), value=x+1)
                    _ = ws5.cell(column=1,
                                 row=3+(j+1)*numBlocks-numBlocks+(x+1), value=j+1)
                    _ = ws5.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                 row=3, value='Scenario:'+str(i+1) )
                    _ = ws5.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                 row=2, value='Area:'+str(k+1) )
                genDFinal[j][k][i] = valueD

    ws5['B3'] = 'block'; ws5['A3'] = 'Stage'

    ###########################################################################

#    ws6 = wb.create_sheet(title="SpillWind")

#    plants = len(windPlants);
#    for i in range(scenarios):
#        for j in range(stages):
#           for k in range(plants):
#                for x in range(blk):
#                    _ = ws6.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
#                                 row=3+(j+1)*blk-blk+(x+1), value=spillWind[i][j][k][x])
#                    _ = ws6.cell(column=2,
#                                 row=3+(j+1)*blk-blk+(x+1), value=x+1)
#                    _ = ws6.cell(column=1,
#                                 row=3+(j+1)*blk-blk+(x+1), value=j+1)
#                    _ = ws6.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
#                                 row=2, value=windPlants[k])
#                    _ = ws6.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
#                                 row=3, value='Scenario:'+str(i+1) )
#    ws6['B3'] = 'block'; ws6['A3'] = 'Stage'

    ###########################################################################

    ws8 = wb.create_sheet(title="LevelBatt")

    plants = len(batteries)
    for i in range(scenarios):
        for j in range(stages):
           for k in range(plants):
                _ = ws8.cell(column=1+(i+1)+(k+1)*scenarios-scenarios,
                             row=4+(j+1), value=lvlBatteries[i][j][k])
                _ = ws8.cell(column=1,
                             row=4+(j+1), value=j+1)
                _ = ws8.cell(column=1+(i+1)+(k+1)*scenarios-scenarios,
                             row=2, value=batteries[k])
                _ = ws8.cell(column=1+(i+1)+(k+1)*scenarios-scenarios,
                             row=3, value='Area:'+str(int(battData[9][k])) )
                _ = ws8.cell(column=1+(i+1)+(k+1)*scenarios-scenarios,
                                 row=4, value='Scenario:'+str(i+1) )
    ws8['A4'] = 'Stage'

    ws7 = wb.create_sheet(title="LoadBatt")
    
    genBdisFinal = [[[[] for y in range(scenarios) ] for z in range(plants) ] for x in range(stages)]
    plants = len(batteries)
    for i in range(scenarios):
        for j in range(stages):
            for k in range(plants):
                valueBdis = 0
                for x in range(numBlocks):
                    valueBdis += loadBatteries[i][j][k][x]
                    
                    _ = ws7.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                 row=4+(j+1)*numBlocks-numBlocks+(x+1), value=loadBatteries[i][j][k][x])
                    _ = ws7.cell(column=2,
                                 row=4+(j+1)*numBlocks-numBlocks+(x+1), value=x+1)
                    _ = ws7.cell(column=1,
                                 row=4+(j+1)*numBlocks-numBlocks+(x+1), value=j+1)
                    _ = ws7.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                 row=2, value=batteries[k])
                    _ = ws7.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                 row=3, value='Area:'+str(int(battData[9][k])) )
                    _ = ws7.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                 row=4, value='Scenario:'+str(i+1) )
                genBdisFinal[j][k][i] = valueBdis
                
    ws7['B4'] = 'block'; ws7['A4'] = 'Stage'

    genBdisBlockFinal = [[0]*numBlocks for x in range(stages)]
    for i in range(scenarios):
        for j in range(stages):
            for k in range(plants):
                for x in range(numBlocks):
                    genBdisBlockFinal[j][x] += loadBatteries[i][j][k][x]
                    
    ###########################################################################
    
    lineval = int(len(linesData))
    transFinal = [[] for y in range(lineval) ]
    
    if numAreas != 1:

        scnTransfer = [[[[[] for y in lensc] for y in lenstg] for a in range(numAreas)] for z in range(numAreas)]

        ws10 = wb.create_sheet(title="EnergyTransfer")
        for i in range(scenarios):
            _ = ws10.cell(column=4+(i+1),row=2, value='Scenario:'+str(i+1) )
            for j in range(stages):
                for k in range(lineval):
                    org = linesData[k][0]; dest = linesData[k][1]
                    for x in range(numBlocks):

                        scnTransfer[org-1][dest-1][j][i].append(linTransfer[i][j][org-1][dest-1][x])

                        _ = ws10.cell(column=4+(i+1),
                                     row=2+((j+1)*numBlocks-numBlocks+(x+1))*lineval-lineval+(k+1), value=linTransfer[i][j][org-1][dest-1][x])
                        _ = ws10.cell(column=2,
                                     row=2+((j+1)*numBlocks-numBlocks+(x+1))*lineval-lineval+(k+1), value=x+1)
                        _ = ws10.cell(column=1,
                                     row=2+((j+1)*numBlocks-numBlocks+(x+1))*lineval-lineval+(k+1), value=j+1)
                        _ = ws10.cell(column=3,
                                     row=2+((j+1)*numBlocks-numBlocks+(x+1))*lineval-lineval+(k+1), value=org)
                        _ = ws10.cell(column=4,
                                     row=2+((j+1)*numBlocks-numBlocks+(x+1))*lineval-lineval+(k+1), value=dest)


        ws10['B2'] = 'block'; ws10['A2'] = 'Stage'; ws10['C2'] = 'Area:From'; ws10['D2'] = 'Area:To'

        ws10_1 = wb.create_sheet(title="EnergyTransferStage")
        for k in range(lineval):
            org = linesData[k][0]; dest = linesData[k][1]
            _ = ws10_1.cell(column=1+(k+1), row=2, value='from:'+str(areasData[0][org-1])+' to:'+str(areasData[0][dest-1]))
            for j in range(stages):
                trf_stg = 0
                for i in range(scenarios):
                    trf_stg += sum(scnTransfer[org-1][dest-1][j][i]) 

                _ = ws10_1.cell(column=1+(k+1), row=2+(j+1), value = trf_stg/scenarios)
                _ = ws10_1.cell(column=1, row=2+(j+1), value=j+1)
                transFinal[k].append(trf_stg/scenarios)

        ws10_1['A2'] = 'stage'

        ws10_2 = wb.create_sheet(title="Transferlimits")
        for k in range(lineval):
            org = linesData[k][0]; dest = linesData[k][1]
            _ = ws10_2.cell(column=1+(k+1), row=2, value='from:'+str(areasData[0][org-1])+' to:'+str(areasData[0][dest-1]))
            for j in range(stages):

                _ = ws10_2.cell(column=1+(k+1), row=2+(j+1), value = l_limits[j][org-1][dest-1])
                _ = ws10_2.cell(column=1, row=2+(j+1), value=j+1)

        ws10_2['A2'] = 'stage'

    ###########################################################################

    emssiFinal = [[[] for y in range(scenarios) ] for x in range(stages)]
    
    if Param.emss_curve is True:
        
        ws10_3 = wb.create_sheet(title="Emissions")
    
        for i in range(scenarios):
            for j in range(stages):
                valueD = 0

                for x in range(numBlocks):
                    valueD += emsscurve[i][j][x]

                    _ = ws10_3.cell(column=2+(i+1),
                                 row=3+(j+1)*numBlocks-numBlocks+(x+1),
                                 value = emsscurve[i][j][x])
                    _ = ws10_3.cell(column=2,
                                 row=3+(j+1)*numBlocks-numBlocks+(x+1), value=x+1)
                    _ = ws10_3.cell(column=1,
                                 row = 3+(j+1)*numBlocks-numBlocks+(x+1), value=j+1)
                    _ = ws10_3.cell(column = 2+(i+1), row=3, value='Scenario:'+str(i+1) )
                emssiFinal[j][i] = valueD
    
        ws10_3['B3'] = 'block'; ws10_3['A3'] = 'Stage'
    
    ###########################################################################

    # Renewables generation by areas
    genRnFinal = [[[[] for y in range(scenarios) ] for z in range(numAreas) ] for x in range(stages)]

    if Param.short_term is False:
            
        ws3 = wb.create_sheet(title="RnwsGen")
    
        for i in range(scenarios):
            for j in range(stages):
               for k in range(numAreas):
                    valueW = 0
                    for x in range(numBlocks):
    
                        valueW += genRnws[i][j][k][x]
    
                        _ = ws3.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                     row=3+(j+1)*numBlocks-numBlocks+(x+1),
                                     value=genRnws[i][j][k][x])
                        _ = ws3.cell(column=2,
                                     row=3+(j+1)*numBlocks-numBlocks+(x+1), value=x+1)
                        _ = ws3.cell(column=1,
                                     row=3+(j+1)*numBlocks-numBlocks+(x+1), value=j+1)
                        _ = ws3.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                     row=2, value='Area: '+areasData[0][k])
                        _ = ws3.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                     row=3, value='Scenario:'+str(i+1) )
                    genRnFinal[j][k][i] = valueW
    
        ws3['B3'] = 'block'; ws3['A3'] = 'Stage'
    else:    
        if Param.wind_aprox is True:
            
            ws3 = wb.create_sheet(title="RnwsGen")
        
            for i in range(scenarios):
                for j in range(stages):
                   for k in range(numAreas):
                        valueW = 0
                        for x in range(numBlocks):
        
                            valueW += genwind[i][j][k][x]
        
                            _ = ws3.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                         row=3+(j+1)*numBlocks-numBlocks+(x+1),
                                         value=genwind[i][j][k][x])
                            _ = ws3.cell(column=2,
                                         row=3+(j+1)*numBlocks-numBlocks+(x+1), value=x+1)
                            _ = ws3.cell(column=1,
                                         row=3+(j+1)*numBlocks-numBlocks+(x+1), value=j+1)
                            _ = ws3.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                         row=2, value='Area: '+areasData[0][k])
                            _ = ws3.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                         row=3, value='Scenario:'+str(i+1) )
                        genRnFinal[j][k][i] = valueW
        
            ws3['B3'] = 'block'; ws3['A3'] = 'Stage'
            
        elif Param.dist_f[0] is True or Param.dist_f[1] is True: 
            
            ws10_4 = wb.create_sheet(title="RnwsGen")
        
            for i in range(scenarios):
                for j in range(stages):
                   for k in range(numAreas):
                        valueRn = 0
                        for x in range(numBlocks):
        
                            valueRn += (df_demand[k][j][x] - genRnws[i][j][k][x])
        
                            _ = ws10_4.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                         row=3+(j+1)*numBlocks-numBlocks+(x+1),
                                         value = df_demand[k][j][x] - genRnws[i][j][k][x])
                            _ = ws10_4.cell(column = 2,
                                         row=3+(j+1)*numBlocks-numBlocks+(x+1), value=x+1)
                            _ = ws10_4.cell(column=1,
                                         row=3+(j+1)*numBlocks-numBlocks+(x+1), value=j+1)
                            _ = ws10_4.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                         row=2, value='Area: '+areasData[0][k])
                            _ = ws10_4.cell(column=2+(i+1)+(k+1)*scenarios-scenarios,
                                         row=3, value='Scenario:'+str(i+1) )
                        genRnFinal[j][k][i] = valueRn
        
            ws10_4['B3'] = 'block'; ws10_4['A3'] = 'Stage'
        
    ###########################################################################

#    ws9 = wb.create_sheet(title="OperativeCost")
#
#    for i in range(scenarios):
#        for j in range(stages):
#                _ = ws9.cell(column=1+(i+1),
#                             row=2+(j+1), value=sol_scn[i][j])
#                _ = ws9.cell(column=1,
#                             row=2+(j+1), value=j+1)
#                _ = ws9.cell(column=1+(i+1),
#                                 row=2, value='Scenario:'+str(i+1) )
#    ws9['A2'] = 'Stage'


    wb.save(filename = dest_filename)

    ###########################################################################

    datasol = {"genHFinal":genHFinal,"genTFinal":genTFinal,"genRnFinal":genRnFinal,
               "genDFinal":genDFinal,"genBFinal":genBFinal,"genMFinal":genMFinal,
               "genSFinal":genSFinal,"emssiFinal":emssiFinal,"genBdisFinal":genBdisFinal,
               "genBBlockFinal":genBBlockFinal,"genBdisBlockFinal":genBdisBlockFinal,
               "transFinal":transFinal}

    pickle.dump(datasol, open( "savedata/html_save.p", "wb" ) )

###############################################################################

def xlsfileCon(operative_cost):
    
    from openpyxl import Workbook
    
    # Create file
    wb = Workbook()
    dest_filename = 'results/Convergence.xlsx'

    ws0 = wb.active
    ws0.title = "Upper-Lower Bound"
        
    for i in range(len(operative_cost[0])):
        _ = ws0.cell(column=1,row=2+(i+1), value= i+1)
        _ = ws0.cell(column=2,row=2+(i+1), value= operative_cost[0][i])
        _ = ws0.cell(column=3,row=2+(i+1), value= operative_cost[1][i])
    ws0['B2'] = 'Upper bound'
    ws0['C2'] = 'Lower bound'

    wb.save(filename = dest_filename)

    ###########################################################################

    # datasol = {"genHFinal":genHFinal}

    # pickle.dump(datasol, open( "savedata/html_save_con.p", "wb" ) )